import tkinter as tk
from tkinter import ttk
from openpyxl import *
from tkinter import filedialog
from openpyxl import load_workbook
import pandas as pd
import os

box= []


global selected_file_path
selected_file_path = ""

global new_file_name 
new_file_name = ""
global newpath

global old_file_name



def set_filename(name):
    global selected_file_path
    global old_file_name
    selected_file_path = name
    temp = selected_file_path.rsplit('/', 1)
    old_file_name = temp[1]


def new_save_cordinents():
    global selected_file_path
    global newpath
    global old_file_name
    newpath = selected_file_path.rsplit('/', 1)
    old_file_name = newpath[1]
    return newpath[0] + "/"

def save():
    global new_file_name 
    new_file_name = pop_name_entry.get()
    temp = new_save_cordinents()
    new_file_name = temp + new_file_name+'.xlsx'
    print(new_file_name+'.xlsx')
    
def starting_row():
    return 22

def remove_junk_value(all_val):
    fress_box = []
    for val in all_val:
        if str(val)[0].isdigit():
            #debugging
            #print(val)
            fress_box.append(val)           
    return fress_box

def get_total_student_count(df2):
    for i in range(len(df2)):
            if str(df2[i]).strip().isdigit():
                ts_count = int(df2[i])
    return ts_count

#takes list as input and returns last row number
def get_last_row (df2):
    ts_count = get_total_student_count(df2)
    return 22+int(ts_count)

def get_all_student_id(df2):
    counter = 0 
    student_id = remove_junk_value(df2)
    for student_id in student_id:
        counter+=1
        print(str(counter) + " " +str(student_id))

def print_name_id_serial_ (df):
    df2 = list(df["SL"])
    
    start = starting_row()
    ending = get_total_student_count(df2)
    print("starting row: ", start)
    print("ending row: ", ending)
    for i in range(0, ending):
        temp_lis = list(df[['SL' , 'Name', 'ID']].iloc[i])
        print(temp_lis)
        box.append(temp_lis)

#gets the exect value from the cell
def get_num(filename, column="O", row=21):
    """Read a single cell value from an Excel file"""
    return pd.read_excel(filename, skiprows=row - 1, usecols=column, nrows=1, header=None, names=["Value"]).iloc[0]["Value"]

global total_student_data 

total_student_data = []

#getting the sum of a student 
def student_sum(filename, row=23, total=0):
    global total_student_data
    
    student_got_up_half = [0,0,0,0,0]
    co_index = ['E','F','G','H','J','K','L','M']
    po_index = ['E','F','G','H','J','K','L','M']
    co1_fillmark = get_num(filename, 'U', 21)
    co2_fillmark = get_num(filename, 'V', 21)
    co3_fillmark = get_num(filename, 'W', 21)
    co4_fillmark = get_num(filename, 'X', 21)
    co5_fillmark = get_num(filename, 'Y', 21)
    
    #co_index = ['E','F','G','H','J','K','L']
    sum = 0
    df = pd.read_excel(filename, skiprows=21)
    wb = load_workbook(filename)
    ws = wb.worksheets[0]
    status["text"] = "Processing : " + "Data"  
    for j in range(total):
        temp = []
        co1 = get_num(filename,co_index[0], row)
        ws.cell(row, column=16).value = co1
        ws.cell(row, column=21).value = co1
        temp.append(co1)
        
        co1_percent = co1/co1_fillmark
        print("co1: ", co1_percent)
        ws.cell(row, column=26).value = co1_percent
        ws.cell(row, column=31).value = co1_percent
        if(co1_percent >= 0.5):
            ws.cell(row, column=36).value = "Yes"
            ws.cell(row, column=41).value = "Yes"
            student_got_up_half[0] += 1
        else:
            ws.cell(row, column=36).value = "No"
            ws.cell(row, column=41).value = "No"
        
        
        co2 = get_num(filename,co_index[2], row)
        ws.cell(row, column=17).value = co2
        ws.cell(row, column=22).value = co2
        temp.append(co2)
        
        co2_percent = co2/co2_fillmark
        print("co2: ", co2_percent)
        ws.cell(row, column=27).value = co2_percent
        ws.cell(row, column=32).value = co2_percent
        if(co2_percent >= 0.5):
            ws.cell(row, column=37).value = "Yes"
            ws.cell(row, column=42).value = "Yes"
            student_got_up_half[1] += 1
        else:
            ws.cell(row, column=37).value = "No"
            ws.cell(row, column=42).value = "No"
        
        
        co3 = get_num(filename,co_index[1], row)
        co3 += get_num(filename,co_index[3], row)
        co3 += get_num(filename,co_index[4], row)
        co3 += get_num(filename,co_index[5], row)
        ws.cell(row, column=18).value = co3
        ws.cell(row, column=23).value = co3
        temp.append(co3)
        
        co3_percent = co3/co3_fillmark
        print("co3: ", co3_percent)
        ws.cell(row, column=28).value = co3_percent
        ws.cell(row, column=33).value = co3_percent
        if(co3_percent >= 0.5):
            ws.cell(row, column=38).value = "Yes"
            ws.cell(row, column=43).value = "Yes"
            student_got_up_half[2] += 1
        else:
            ws.cell(row, column=38).value = "No"
            ws.cell(row, column=43).value = "No"
        
        co4 = get_num(filename,co_index[6], row)
        ws.cell(row, column=19).value = co4
        ws.cell(row, column=24).value = co4
        temp.append(co4)
        
        co4_percent = co4/co4_fillmark
        print("co4: ", co4_percent)
        ws.cell(row, column=29).value = co4_percent
        ws.cell(row, column=34).value = co4_percent
        if(co4_percent >= 0.5):
            ws.cell(row, column=39).value = "Yes"
            ws.cell(row, column=44).value = "Yes"
            student_got_up_half[3] += 1
        else:
            ws.cell(row, column=39).value = "No"
            ws.cell(row, column=44).value = "No"
        
        
        co5 = get_num(filename,co_index[7], row)
        ws.cell(row, column=20).value = co5
        ws.cell(row, column=25).value = co5
        temp.append(co5)
        
        co5_percent = co5/co5_fillmark
        print("co5: ", co5_percent)
        ws.cell(row, column=30).value = co5_percent
        ws.cell(row, column=35).value = co5_percent
        if(co5_percent >= 0.5):
            ws.cell(row, column=40).value = "Yes"
            ws.cell(row, column=45).value = "Yes"
            student_got_up_half[4] += 1
        else:
            ws.cell(row, column=40).value = "No"
            ws.cell(row, column=45).value = "No"
        
        
        
        #for i in range(8):
            #print(get_num(filename, co_index[i], row))
            #sum += get_num(filename, co_index[i], row)
        sum = co1 + co2 + co3 + co4 + co5    
        ws.cell(row, column=15).value = sum
            
        print(f"{box[j]} got numbers in total  {j}: ", sum)
        print(temp)
        #total_student_data.append(temp)
        row= row+1
        sum = 0
        status["text"] = "Done: " + str(j+1) + " of " + str(total) + " Data"
        status.update()
    
    for i in range(5,16):
        if(i!=10):
            ws.cell(271, column=i).value = total
    
    
    
    counter = 0
    for i in range(5,16):
        if(i!=10):
            if(counter == 5):
                counter = 0
            ws.cell(273, column=i).value = student_got_up_half[counter]
            ws.cell(274, column=i).value = student_got_up_half[counter]/total
            counter += 1 
    counter = 0
    for i in range(5,10):
        ws.cell(283, column=i).value = student_got_up_half[counter]
        ws.cell(284, column=i).value = student_got_up_half[counter]/total
        if(student_got_up_half[counter]/total>=0.5):
            ws.cell(285, column=i).value = "Yes"
        else:
            ws.cell(285, column=i).value = "No"
        
        ws.cell(287, column=i).value = student_got_up_half[counter]
        ws.cell(288, column=i).value = student_got_up_half[counter]/total
        if(student_got_up_half[counter]/total>=0.5):
            ws.cell(289, column=i).value = "Yes"
        else:
            ws.cell(289, column=i).value = "No"
        counter += 1
    
    
    percentage_of_student_who_got = []
    
    
    ws.cell(281, column=2).value = "Total Number of student in this courser: " + str(total)
    
    global new_file_name
    global old_file_name
    if(new_file_name==""):
        wb.save(filename)
    else:
        wb.save(new_file_name)
    
    
    print("student_got_up_half: ", student_got_up_half)
    status["text"] = "task :" + "Done"


#handles file selection and file reading and branching
def openfile():
    filename = filedialog.askopenfilename(
        initialdir="/",
        title="Select CSV",
        filetypes=(("xlsx files", "*.xlsx"),("csv files", "*.csv")),
        # filetypes=(("csv files", "*.csv"),("xlsx files", "*.xlsx")),
    )
    print(filename)
    if filename == "":
        status["text"] = "No File Selected"
    else:
        set_filename(filename)
        status["text"] = "File Selected"
        # df = pd.read_excel(filename, skiprows=21)
        # print_name_id_serial_(df)
        # df2 = list(df["SL"])
        # #get_last_row(df2)
        # total = get_total_student_count(df2)
        
        
        
        # print("total student: ", get_total_student_count(df2))
        
        # student_sum(filename, 23, total)
        #status["text"] = "Total Student: " + str(ts_count)
        #print("last_row number: ", get_last_row(df2))
        
        #student_id =  list(df["ID"])
        #get_all_student_id(student_id)
        
            
#debugging
def submit():
    global selected_file_path
    global total_student_data
    if(selected_file_path == ""):
        status["text"] = "please select a file"
    else:
        status["text"] = "File Selected Successfully"
        df = pd.read_excel(selected_file_path, skiprows=21)
        print_name_id_serial_(df)
        df2 = list(df["SL"])
        #get_last_row(df2)
        total = get_total_student_count(df2)
        print("total student: ", get_total_student_count(df2))
        
        student_sum(selected_file_path, 23, total)
        print("->file name is  : ", selected_file_path)
        #print(total_student_data)


def center(win):
    """
    centers a tkinter window
    :param win: the main window or Toplevel window to center
    """
    win.update_idletasks()
    width = win.winfo_width()
    frm_width = win.winfo_rootx() - win.winfo_x()
    win_width = width + 2 * frm_width
    height = win.winfo_height()
    titlebar_height = win.winfo_rooty() - win.winfo_y()
    win_height = height + titlebar_height + frm_width
    x = win.winfo_screenwidth() // 2 - win_width // 2
    y = win.winfo_screenheight() // 2 - win_height // 2
    win.geometry('{}x{}+{}+{}'.format(width, height, x, y))
    win.deiconify()



#main window
window = tk.Tk()

window.iconbitmap("icon.ico")

#clears the terminal
os.system("cls")
#window title
window.title("CoPoSystem")
#initial window size
window.geometry("400x130")
#globally sets the font
window.option_add("*font", "Helvetica 8 bold")



main_frame = tk.Frame(window, background='white')
main_frame.pack(fill="both", expand=True)



s = ttk.Style()
s.configure('TButton', font=('Helvetica', 8, 'bold') ,background='#7AC5CD')



submit_btn = ttk.Button(main_frame, width=30, text="Submit" ,command=submit  )
submit_btn.grid( row=0,column=0, padx=5, pady=10 ,sticky='W')

file_selecter = ttk.Button(main_frame, width=30, text="Select File", command=openfile)
file_selecter.grid(row=0, column=1, padx=5, pady=10 ,sticky='W')

def popup_bonus():
    win = tk.Toplevel()
    win.wm_title("Insert A File Name : ")
    win.geometry("300x130")
    win.iconbitmap("icon.ico")
    
    
    global pop_name_entry
    pop_name_entry = ttk.Entry(win, font=("default", (12)),width=30)
    pop_name_entry.pack(padx=5, pady=10)

    
    pop = ttk.Button(win, text="Submit",width=40, command=win.destroy )
    pop.pack(padx=5, pady=10)
    save_name = ttk.Button(win, text="save",width=40, command=save )
    save_name.pack(padx=5, pady=10)
    center(win)
    

name_box = ttk.Button(main_frame, width=30, text="File name" , command = popup_bonus)
name_box.grid(row=1, column=1, padx=5, pady=10 ,sticky='W')


status_lable = ttk.LabelFrame(main_frame, width=30, text="Process : "  ,style='new.TFrame')
status_lable.grid(row=1 , padx=5, pady=10, sticky='NWES')

status = ttk.Label(status_lable, text="idle")
status.grid(padx=5, pady=10, sticky="NWES")

center(window)


window.mainloop()
